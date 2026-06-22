"""
Excel Validator

Responsibilities:
- Validate Excel structure
- Validate required sheets
- Validate required columns
- Validate duplicate template keys
- Validate duplicate pattern IDs
- Validate template syntax
- Validate enabled flags
- Validate priority values
"""

import io
from pathlib import Path
from typing import Dict, List, Set

from openpyxl import load_workbook

from utils import setup_logger

logger = setup_logger(__name__)

SHEET_NAME = "step_patterns"

REQUIRED_COLUMNS = {
    "pattern_id",
    "action",
    "instruction_pattern",
    "description",
    "example_1",
    "priority",
}

VALID_ENABLED_VALUES = {
    "true",
    "false",
    "1",
    "0",
    "yes",
    "no",
}

VALID_ACTIONS = {
    "click",
    "enter",
    "select",
    "verify",
    "navigate",
    "memory",
    "memory_store",
    "generate",
    "upload",
    "capture",
    "clear",
}


class DynamicExcelValidator:
    """
    Dynamic Excel validation engine.
    """

    def __init__(
        self,
        excel_path: Path,
    ):

        self.excel_path = excel_path

    def validate(
        self,
    ) -> bool:
        """
        Validate dynamic pattern Excel from a file path.

        Returns:
            bool
        """

        if not self.excel_path.exists():

            raise FileNotFoundError(
                f"Excel file not found: "
                f"{self.excel_path}"
            )

        workbook = load_workbook(
            filename=self.excel_path,
            data_only=True,
        )

        return self._validate_workbook(workbook)

    def validate_from_bytes(
        self,
        data: io.BytesIO,
    ) -> bool:
        """
        Validate dynamic pattern Excel from an in-memory BytesIO buffer.
        No file path needed — zero disk I/O.

        Returns:
            bool
        """

        data.seek(0)

        workbook = load_workbook(
            filename=data,
            data_only=True,
        )

        return self._validate_workbook(workbook)

    def _validate_workbook(
        self,
        workbook,
    ) -> bool:
        """Shared validation logic for both file-path and in-memory paths."""

        self._validate_sheet_exists(
            workbook,
        )

        sheet = workbook[SHEET_NAME]

        headers = self._extract_headers(
            sheet,
        )

        self._validate_required_columns(
            headers,
        )

        self._validate_rows(
            sheet=sheet,
            headers=headers,
        )

        logger.info(
            "Dynamic Excel validation successful"
        )

        return True

    def _validate_sheet_exists(
        self,
        workbook,
    ) -> None:

        if SHEET_NAME not in workbook.sheetnames:

            raise ValueError(
                f"Required sheet '{SHEET_NAME}' not found"
            )

    def _extract_headers(
        self,
        sheet,
    ) -> Dict[str, int]:

        headers = {}

        for index, cell in enumerate(
            sheet[1]
        ):

            if cell.value is None:
                continue

            header = (
                str(cell.value)
                .strip()
                .lower()
            )

            headers[header] = index

        return headers

    def _validate_required_columns(
        self,
        headers: Dict[str, int],
    ) -> None:

        missing_columns = (
            REQUIRED_COLUMNS - set(headers.keys())
        )

        if missing_columns:

            raise ValueError(
                "Missing required columns: "
                f"{sorted(missing_columns)}"
            )

    def _validate_rows(
        self,
        sheet,
        headers: Dict[str, int],
    ) -> None:

        pattern_ids_seen: Set[str] = set()

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2),
            start=2,
        ):

            row_data = self._extract_row_data(
                row=row,
                headers=headers,
            )

            if not any(row_data.values()):
                continue

            self._validate_pattern_id(
                row_data,
                pattern_ids_seen,
                row_index,
            )

            self._validate_category(
                row_data,
                row_index,
            )

            self._validate_action(
                row_data,
                row_index,
            )

            self._validate_instruction_pattern(
                row_data,
                row_index,
            )

            if "enabled" in row_data:
                self._validate_enabled(
                    row_data,
                    row_index,
                )

            self._validate_priority(
                row_data,
                row_index,
            )

    def _extract_row_data(
        self,
        row,
        headers: Dict[str, int],
    ) -> Dict[str, str]:

        row_data = {}

        for header, index in headers.items():

            cell_value = row[index].value

            row_data[header] = (
                ""
                if cell_value is None
                else str(cell_value).strip()
            )

        return row_data

    def _validate_pattern_id(
        self,
        row_data: Dict[str, str],
        pattern_ids_seen: Set[str],
        row_index: int,
    ) -> None:

        pattern_id = row_data.get(
            "pattern_id",
            "",
        )

        if not pattern_id:

            raise ValueError(
                f"Empty pattern_id at row {row_index}"
            )

        if pattern_id in pattern_ids_seen:

            raise ValueError(
                f"Duplicate pattern_id "
                f"'{pattern_id}' at row {row_index}"
            )

        pattern_ids_seen.add(
            pattern_id
        )

    def _validate_category(
        self,
        row_data: Dict[str, str],
        row_index: int,
    ) -> None:

        category = (
            row_data.get("template_key")
            or row_data.get("category")
        )

        if not category:

            raise ValueError(
                f"Missing category or template_key at row {row_index}"
            )

    def _validate_action(
        self,
        row_data: Dict[str, str],
        row_index: int,
    ) -> None:

        action = (
            row_data.get(
                "action",
                "",
            )
            .strip()
            .lower()
        )

        if not action:

            raise ValueError(
                f"Empty action at row {row_index}"
            )

    def _validate_instruction_pattern(
        self,
        row_data: Dict[str, str],
        row_index: int,
    ) -> None:

        instruction_pattern = row_data.get(
            "instruction_pattern",
            "",
        )

        if not instruction_pattern:

            raise ValueError(
                f"Empty instruction_pattern at row {row_index}"
            )

        if (
            instruction_pattern.count("<<")
            != instruction_pattern.count(">>")
        ):

            raise ValueError(
                f"Invalid placeholder syntax "
                f"at row {row_index}"
            )

        if len(instruction_pattern.strip()) < 5:

            raise ValueError(
                f"Instruction pattern too short at row "
                f"{row_index}"
            )

    def _validate_enabled(
        self,
        row_data: Dict[str, str],
        row_index: int,
    ) -> None:

        enabled = (
            row_data.get(
                "enabled",
                "",
            )
            .strip()
            .lower()
        )

        if enabled not in VALID_ENABLED_VALUES:

            raise ValueError(
                f"Invalid enabled value "
                f"'{enabled}' at row {row_index}"
            )

    def _validate_priority(
        self,
        row_data: Dict[str, str],
        row_index: int,
    ) -> None:

        priority = row_data.get(
            "priority",
            "",
        )

        try:

            # Google Sheets exports numbers as floats (e.g. "1.0").
            # Convert via float() first so both "1" and "1.0" are accepted.
            priority_value = int(float(priority))

            if priority_value < 1:

                raise ValueError

        except Exception:

            raise ValueError(
                f"Invalid priority value "
                f"'{priority}' at row {row_index}"
            )


def validate_dynamic_excel(
    excel_path: Path,
) -> bool:
    """
    Validate dynamic Excel file from disk path.

    Returns:
        bool
    """

    validator = DynamicExcelValidator(
        excel_path=excel_path,
    )

    return validator.validate()


def validate_dynamic_excel_from_bytes(
    data: io.BytesIO,
) -> bool:
    """
    Validate dynamic Excel from an in-memory BytesIO buffer.
    No disk path required — zero disk I/O.

    Returns:
        bool
    """

    # excel_path is unused when calling validate_from_bytes,
    # but DynamicExcelValidator.__init__ requires it.
    dummy_path = Path("in_memory.xlsx")

    validator = DynamicExcelValidator(
        excel_path=dummy_path,
    )

    return validator.validate_from_bytes(data)