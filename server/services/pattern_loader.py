"""
Pattern Loader

Responsibilities:
- Load templates from Excel
- Validate required columns
- Skip disabled rows
- Normalize template structure
- Store templates in runtime cache
- Provide safe fallback support
"""

import io
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from config.config_loader import get_config
from utils import setup_logger

logger = setup_logger(__name__)

REQUIRED_COLUMNS = {
    "pattern_id",
    "action",
    "instruction_pattern",
    "description",
    "example_1",
    "priority",
}

SHEET_NAME = "step_patterns"


class DynamicPatternLoader:
    """
    Dynamic Excel template loader.
    """

    def __init__(
        self,
        excel_path: Path | None = None,
    ):

        self.config = get_config()

        self.excel_path = (
            excel_path
            or self.config.get_pattern_excel_path()
        )

    def load_patterns(
        self,
    ) -> Dict[str, Dict[str, Any]]:
        """
        Load dynamic templates from an Excel file on disk.

        Returns:
            {"button_click": {...}}
        """

        if not self.excel_path.exists():

            logger.warning(
                "Dynamic pattern Excel not found | %s",
                self.excel_path,
            )

            return {}

        try:

            workbook = load_workbook(
                filename=self.excel_path,
                data_only=True,
            )

            return self._parse_workbook(workbook)

        except Exception:

            logger.exception(
                "Failed loading dynamic patterns"
            )

            return {}

    def load_patterns_from_bytes(
        self,
        data: io.BytesIO,
    ) -> Dict[str, Dict[str, Any]]:
        """
        Load dynamic templates from an in-memory BytesIO buffer.
        No file is read from disk - zero disk I/O.

        Returns:
            {"button_click": {...}}
        """

        try:

            data.seek(0)

            workbook = load_workbook(
                filename=data,
                data_only=True,
            )

            return self._parse_workbook(workbook)

        except Exception:

            logger.exception(
                "Failed loading dynamic patterns from bytes"
            )

            return {}

    def _parse_workbook(
        self,
        workbook,
    ) -> Dict[str, Dict[str, Any]]:
        """Shared parse logic for both disk and in-memory paths."""

        if SHEET_NAME not in workbook.sheetnames:

            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found"
            )

        sheet = workbook[SHEET_NAME]

        headers = self._extract_headers(
            sheet,
        )

        self._validate_headers(
            headers,
        )

        patterns = self._parse_sheet(
            sheet=sheet,
            headers=headers,
        )

        logger.info(
            "Dynamic patterns loaded successfully | count=%s",
            len(patterns),
        )

        return patterns

    def _extract_headers(
        self,
        sheet,
    ) -> Dict[str, int]:

        headers = {}

        for index, cell in enumerate(
            sheet[1]
        ):

            if cell.value is None:
                continue

            header_name = (
                str(cell.value)
                .strip()
                .lower()
            )

            headers[header_name] = index

        return headers

    def _validate_headers(
        self,
        headers: Dict[str, int],
    ) -> None:

        missing_columns = (
            REQUIRED_COLUMNS - set(headers.keys())
        )

        if missing_columns:

            raise ValueError(
                "Missing required Excel columns: "
                f"{sorted(missing_columns)}"
            )

    def _parse_sheet(
        self,
        sheet,
        headers: Dict[str, int],
    ) -> Dict[str, Dict[str, Any]]:

        patterns = {}

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2),
            start=2,
        ):

            try:

                row_data = self._extract_row_data(
                    row=row,
                    headers=headers,
                )

                if not row_data:
                    continue

                enabled = (
                    str(
                        row_data.get(
                            "enabled",
                            "true",
                        )
                    )
                    .strip()
                    .lower()
                )

                if enabled not in {
                    "true",
                    "1",
                    "yes",
                    "",
                }:
                    continue

                category = (
                    row_data.get("template_key")
                    or row_data.get("category")
                )

                if not category:

                    raise ValueError(
                        f"Missing category or template_key "
                        f"at row {row_index}"
                    )

                category = str(category).strip()

                category = str(category).strip()

                instruction_pattern = (
                    row_data["instruction_pattern"]
                    .strip()
                )

                self._validate_instruction_pattern(
                    instruction_pattern=instruction_pattern,
                    row_index=row_index,
                )

                examples = []

                example_1 = (
                    str(
                        row_data.get(
                            "example_1",
                            "",
                        )
                    ).strip()
                )

                example_2 = (
                    str(
                        row_data.get(
                            "example_2",
                            "",
                        )
                    ).strip()
                )

                if example_1:
                    examples.append(example_1)

                if example_2:
                    examples.append(example_2)

                pattern_id = str(row_data["pattern_id"]).strip()

                patterns[pattern_id] = {
                    "pattern_id":
                        pattern_id,

                    "template_key":
                        category,

                    "action":
                        str(
                            row_data["action"]
                        ).strip(),

                    "template":
                        instruction_pattern,

                    "description":
                        str(
                            row_data["description"]
                        ).strip(),

                    "examples":
                        examples,

                    "priority":
                        int(
                            float(
                                row_data.get(
                                    "priority",
                                    1,
                                )
                            )
                        ),
                }

            except Exception:

                logger.exception(
                    "Failed parsing row | row=%s",
                    row_index,
                )

        return patterns

    def _extract_row_data(
        self,
        row,
        headers: Dict[str, int],
    ) -> Dict[str, Any]:

        row_data = {}

        for header, index in headers.items():

            cell_value = row[index].value

            row_data[header] = (
                ""
                if cell_value is None
                else str(cell_value)
            )

        return row_data

    def _validate_instruction_pattern(
        self,
        instruction_pattern: str,
        row_index: int,
    ) -> None:

        if not instruction_pattern:

            raise ValueError(
                f"Empty instruction_pattern at row "
                f"{row_index}"
            )

        if (
            instruction_pattern.count("<<")
            != instruction_pattern.count(">>")
        ):

            raise ValueError(
                f"Invalid placeholder syntax "
                f"at row {row_index}"
            )

    def register_patterns(
        self,
    ) -> Dict[str, Dict[str, Any]]:
        """Register patterns loaded from disk into the runtime cache."""

        patterns = self.load_patterns()

        return self._register(patterns)

    def register_patterns_from_bytes(
        self,
        data: io.BytesIO,
    ) -> Dict[str, Dict[str, Any]]:
        """
        Register patterns loaded from an in-memory BytesIO into the runtime cache.
        Zero disk I/O - preferred in production.
        """

        patterns = self.load_patterns_from_bytes(data)

        return self._register(patterns)

    def _register(
        self,
        patterns: Dict[str, Dict[str, Any]],
    ) -> Dict[str, Dict[str, Any]]:
        """Shared cache-registration logic."""

        if patterns:

            self.config.set_runtime_cache(
                "dynamic_templates",
                patterns,
            )

            logger.info(
                "Dynamic templates registered in runtime cache"
            )

        else:

            logger.warning(
                "No dynamic templates registered"
            )

        return patterns


def init_patterns() -> Dict[str, Dict[str, Any]]:
    """
    Initialize dynamic patterns from disk (startup / fallback).
    """

    loader = DynamicPatternLoader()

    return loader.register_patterns()


def init_patterns_from_bytes(
    data: io.BytesIO,
) -> Dict[str, Dict[str, Any]]:
    """
    Initialize dynamic patterns from an in-memory BytesIO buffer.
    Zero disk I/O - preferred in production.
    """

    loader = DynamicPatternLoader()

    return loader.register_patterns_from_bytes(data)